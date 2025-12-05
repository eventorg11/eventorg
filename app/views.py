from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.contrib.auth import login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db import IntegrityError, models
from app.models import Conference, EventRegistration, UserProfile, ContactMessage
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

User = get_user_model()
from app.forms import (
    CustomUserCreationForm,
    CustomAuthenticationForm,
    ProfileUpdateForm,
    ConferenceCreationForm,
    ContactForm,
)


def index(request):
    """Главная страница с ближайшими мероприятиями"""
    now = timezone.now()
    
    upcoming_conferences = Conference.objects.filter(
        start_date__gte=now
    ).order_by('start_date')[:2]
    
    if upcoming_conferences.count() < 2:
        additional_needed = 2 - upcoming_conferences.count()
        additional = Conference.objects.filter(
            start_date__lt=now
        ).order_by('-start_date')[:additional_needed]
        upcoming_conferences = list(upcoming_conferences) + list(additional)
    
    return render(request, "index.html", {"upcoming_conferences": upcoming_conferences})


def about(request):
    return render(request, "about.html")


def events(request):
    """Страница со списком всех мероприятий с поиском и фильтрами"""
    conferences = Conference.objects.all()
    
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    type_filter = request.GET.get('type', '')
    sort_by = request.GET.get('sort', '-start_date')
    
    if search_query:
        conferences = conferences.filter(
            models.Q(title__icontains=search_query) |
            models.Q(description__icontains=search_query) |
            models.Q(short_description__icontains=search_query)
        )
    
    if status_filter:
        conferences = conferences.filter(status=status_filter)
    
    if type_filter == 'online':
        conferences = conferences.filter(is_online=True)
    elif type_filter == 'offline':
        conferences = conferences.filter(is_online=False)
    
    valid_sort_options = ['-start_date', 'start_date', '-created_at', 'created_at', 'title', '-title']
    if sort_by not in valid_sort_options:
        sort_by = '-start_date'
    
    conferences = conferences.order_by(sort_by)
    
    context = {
        'conferences': conferences,
        'search_query': search_query,
        'status_filter': status_filter,
        'type_filter': type_filter,
        'sort_by': sort_by,
    }
    
    return render(request, "events.html", context)


def contact(request):
    """Страница контактов"""
    return render(request, "contact.html")


def feedback(request):
    """Страница формы обратной связи"""
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            contact_message = ContactMessage.objects.create(
                name=form.cleaned_data['name'],
                email=form.cleaned_data['email'],
                subject=form.cleaned_data['subject'],
                message=form.cleaned_data['message'],
                user=request.user if request.user.is_authenticated else None,
                status='new'
            )
            context = {
                'form': ContactForm(),
                'success_message': 'Ваше сообщение успешно отправлено! Мы свяжемся с вами в ближайшее время.',
            }
            if request.user.is_authenticated:
                context['form'].fields['name'].initial = request.user.get_full_name() or request.user.username
                context['form'].fields['email'].initial = request.user.email
            return render(request, "feedback.html", context)
    else:
        form = ContactForm()
        if request.user.is_authenticated:
            form.fields['name'].initial = request.user.get_full_name() or request.user.username
            form.fields['email'].initial = request.user.email
    
    context = {
        'form': form,
    }
    return render(request, "feedback.html", context)


def faq(request):
    """Страница часто задаваемых вопросов"""
    return render(request, "faq.html")


def event_detail(request, event_id):
    """Детализированная страница просмотра мероприятия"""
    conference = get_object_or_404(Conference, id=event_id)
    is_registered = False
    is_curator_of_event = False
    
    if request.user.is_authenticated:
        is_registered = EventRegistration.objects.filter(
            user=request.user,
            conference=conference
        ).exists()
        
        user_profile = getattr(request.user, 'profile', None)
        is_curator = user_profile and user_profile.is_curator()
        is_admin = user_profile and user_profile.is_admin()
        
        if is_curator and conference.curator == request.user:
            is_curator_of_event = True
        elif is_admin:
            is_curator_of_event = True
    
    context = {
        'conference': conference,
        'is_registered': is_registered,
        'is_curator_of_event': is_curator_of_event,
    }
    return render(request, "event_detail.html", context)


@login_required
def register_for_event(request, event_id):
    """Регистрация пользователя на мероприятие"""
    conference = get_object_or_404(Conference, id=event_id)
    
    if request.method == 'POST':
        if not conference.is_registration_open:
            return JsonResponse({'success': False, 'error': 'Регистрация на это мероприятие закрыта'}, status=400)
        
        if conference.max_participants:
            current_registrations = EventRegistration.objects.filter(conference=conference).count()
            if current_registrations >= conference.max_participants:
                return JsonResponse({'success': False, 'error': 'Достигнуто максимальное количество участников'}, status=400)
        
        try:
            EventRegistration.objects.create(user=request.user, conference=conference)
            return JsonResponse({'success': True, 'message': 'Вы успешно зарегистрированы на мероприятие'})
        except IntegrityError:
            return JsonResponse({'success': False, 'error': 'Вы уже зарегистрированы на это мероприятие'}, status=400)
    
    return JsonResponse({'success': False, 'error': 'Неверный метод запроса'}, status=405)


def register_view(request):
    """Страница регистрации"""
    if request.user.is_authenticated:
        return redirect('index')
    
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('index')
    else:
        form = CustomUserCreationForm()
    
    return render(request, 'registration/register.html', {'form': form})


def login_view(request):
    """Страница входа"""
    if request.user.is_authenticated:
        return redirect('index')
    
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            next_url = request.GET.get('next', 'index')
            return redirect(next_url)
    else:
        form = CustomAuthenticationForm()
    
    return render(request, 'registration/login.html', {'form': form})


@login_required
def logout_view(request):
    """Выход из системы"""
    logout(request)
    return redirect('index')


@login_required
def profile_view(request):
    """Страница личного кабинета пользователя"""
    user_profile = getattr(request.user, 'profile', None)
    is_curator = user_profile and user_profile.is_curator()
    is_admin = user_profile and user_profile.is_admin()
    
    if is_admin:
        users = User.objects.all().select_related('profile').order_by('username')
        contact_messages = ContactMessage.objects.all().order_by('-created_at')
        context = {
            'is_admin': True,
            'users': users,
            'contact_messages': contact_messages,
        }
    elif is_curator:
        curated_conferences = Conference.objects.filter(
            curator=request.user
        ).order_by('-created_at')
        
        for conference in curated_conferences:
            conference.registration_count = EventRegistration.objects.filter(
                conference=conference
            ).count()
        
        context = {
            'is_admin': False,
            'is_curator': True,
            'curated_conferences': curated_conferences,
        }
    else:
        registrations = EventRegistration.objects.filter(
            user=request.user
        ).select_related('conference').order_by('-created_at')
        context = {
            'is_admin': False,
            'is_curator': False,
            'registrations': registrations,
        }
    
    return render(request, 'profile.html', context)


@login_required
def cancel_registration(request, registration_id):
    """Отмена регистрации пользователя на мероприятие"""
    registration = get_object_or_404(
        EventRegistration,
        id=registration_id,
        user=request.user
    )
    
    if request.method == 'POST':
        conference_title = registration.conference.title
        registration.delete()
        return JsonResponse({
            'success': True,
            'message': f'Регистрация на мероприятие "{conference_title}" отменена'
        })
    
    return JsonResponse({'success': False, 'error': 'Неверный метод запроса'}, status=405)


@login_required
def settings_view(request):
    """Страница настроек пользователя"""
    from django.contrib.auth.forms import PasswordChangeForm

    profile_success = False
    password_success = False

    profile_form = ProfileUpdateForm(instance=request.user)
    password_form = PasswordChangeForm(user=request.user)
    for field in password_form.fields.values():
        field.widget.attrs.update({"class": "form-control"})

    if request.method == 'POST':
        if 'profile_submit' in request.POST:
            profile_form = ProfileUpdateForm(request.POST, instance=request.user)
            if profile_form.is_valid():
                profile_form.save()
                profile_success = True
            password_form = PasswordChangeForm(user=request.user)
            for field in password_form.fields.values():
                field.widget.attrs.update({"class": "form-control"})
        elif 'password_submit' in request.POST:
            password_form = PasswordChangeForm(user=request.user, data=request.POST)
            profile_form = ProfileUpdateForm(instance=request.user)
            for field in password_form.fields.values():
                field.widget.attrs.update({"class": "form-control"})
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                password_success = True

    context = {
        'profile_form': profile_form,
        'password_form': password_form,
        'profile_success': profile_success,
        'password_success': password_success,
    }
    return render(request, 'settings.html', context)


@login_required
def event_participants_view(request, event_id):
    """Страница со списком зарегистрированных пользователей на мероприятие"""
    conference = get_object_or_404(Conference, id=event_id)
    
    user_profile = getattr(request.user, 'profile', None)
    is_curator = user_profile and user_profile.is_curator()
    is_admin = user_profile and user_profile.is_admin()
    
    if not (is_curator and conference.curator == request.user) and not is_admin:
        return redirect('profile')
    
    registrations = EventRegistration.objects.filter(
        conference=conference
    ).select_related('user').order_by('-created_at')
    
    context = {
        'conference': conference,
        'registrations': registrations,
        'registration_count': registrations.count(),
        'max_participants': conference.max_participants,
    }
    
    return render(request, 'event_participants.html', context)


@login_required
def remove_participant(request, registration_id):
    """Удаление регистрации участника (для куратора)"""
    registration = get_object_or_404(EventRegistration, id=registration_id)
    conference = registration.conference
    
    user_profile = getattr(request.user, 'profile', None)
    is_curator = user_profile and user_profile.is_curator()
    is_admin = user_profile and user_profile.is_admin()
    
    if not (is_curator and conference.curator == request.user) and not is_admin:
        return JsonResponse({'success': False, 'error': 'Нет доступа'}, status=403)
    
    if request.method == 'POST':
        user_name = f"{registration.user.first_name} {registration.user.last_name}".strip() or registration.user.username
        registration.delete()
        return JsonResponse({
            'success': True,
            'message': f'Регистрация пользователя {user_name} успешно удалена'
        })
    
    return JsonResponse({'success': False, 'error': 'Неверный метод запроса'}, status=400)


@login_required
def create_event_view(request):
    """Страница создания мероприятия для куратора"""
    user_profile = getattr(request.user, 'profile', None)
    is_curator = user_profile and user_profile.is_curator()
    is_admin = user_profile and user_profile.is_admin()
    
    if not is_curator and not is_admin:
        return redirect('profile')
    
    if request.method == 'POST':
        form = ConferenceCreationForm(request.POST)
        if form.is_valid():
            conference = form.save(commit=False)
            conference.curator = request.user
            conference.status = 'not_started'
            conference.save()
            return redirect('profile')
    else:
        form = ConferenceCreationForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'create_event.html', context)


@login_required
def assign_curator(request):
    """Назначение пользователя куратором (для админа)"""
    user_profile = getattr(request.user, 'profile', None)
    is_admin = user_profile and user_profile.is_admin()
    
    if not is_admin:
        return JsonResponse({'success': False, 'error': 'Нет доступа'}, status=403)
    
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Не указан пользователь'}, status=400)
        
        try:
            user = User.objects.get(id=user_id)
            profile, created = UserProfile.objects.get_or_create(user=user)
            profile.role = 'curator'
            profile.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Пользователь {user.username} успешно назначен куратором'
            })
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Пользователь не найден'}, status=404)
    
    return JsonResponse({'success': False, 'error': 'Неверный метод запроса'}, status=400)


@login_required
def delete_user(request):
    """Удаление пользователя (для админа)"""
    user_profile = getattr(request.user, 'profile', None)
    is_admin = user_profile and user_profile.is_admin()
    
    if not is_admin:
        return JsonResponse({'success': False, 'error': 'Нет доступа'}, status=403)
    
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'Не указан пользователь'}, status=400)
        
        try:
            user = User.objects.get(id=user_id)
            
            if user == request.user:
                return JsonResponse({'success': False, 'error': 'Нельзя удалить самого себя'}, status=400)
            
            username = user.username
            user.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Пользователь {username} успешно удален'
            })
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Пользователь не найден'}, status=404)
    
    return JsonResponse({'success': False, 'error': 'Неверный метод запроса'}, status=400)


@login_required
def export_conferences_excel(request):
    """Экспорт информации о конференциях в Excel файл (только для администраторов)"""
    user_profile = getattr(request.user, 'profile', None)
    is_admin = user_profile and user_profile.is_admin()
    
    if not is_admin:
        return JsonResponse({'success': False, 'error': 'Нет доступа'}, status=403)
    
    conferences = Conference.objects.all().order_by('start_date')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Конференции"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    headers = ['Название конференции', 'Дата начала', 'Количество зарегистрированных', 'Максимум мест']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for conference in conferences:
        registration_count = EventRegistration.objects.filter(conference=conference).count()
        max_participants = conference.max_participants if conference.max_participants else 'Не ограничено'
        start_date = timezone.localtime(conference.start_date).strftime('%d.%m.%Y %H:%M')
        
        ws.append([
            conference.title,
            start_date,
            registration_count,
            max_participants
        ])
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'conferences_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def update_message_status(request, message_id):
    """Обновление статуса сообщения обратной связи (для админа)"""
    user_profile = getattr(request.user, 'profile', None)
    is_admin = user_profile and user_profile.is_admin()
    
    if not is_admin:
        return JsonResponse({'success': False, 'error': 'Нет доступа'}, status=403)
    
    if request.method == 'POST':
        try:
            message = ContactMessage.objects.get(id=message_id)
            new_status = request.POST.get('status')
            
            if new_status in ['new', 'read', 'replied']:
                message.status = new_status
                message.save()
                return JsonResponse({
                    'success': True,
                    'message': 'Статус сообщения обновлен',
                    'new_status': message.get_status_display()
                })
            else:
                return JsonResponse({'success': False, 'error': 'Неверный статус'}, status=400)
        except ContactMessage.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Сообщение не найдено'}, status=404)
    
    return JsonResponse({'success': False, 'error': 'Неверный метод запроса'}, status=400)


@login_required
def delete_message(request, message_id):
    """Удаление сообщения обратной связи (для админа)"""
    user_profile = getattr(request.user, 'profile', None)
    is_admin = user_profile and user_profile.is_admin()
    
    if not is_admin:
        return JsonResponse({'success': False, 'error': 'Нет доступа'}, status=403)
    
    if request.method == 'POST':
        try:
            message = ContactMessage.objects.get(id=message_id)
            subject = message.subject
            message.delete()
            return JsonResponse({
                'success': True,
                'message': f'Сообщение "{subject}" успешно удалено'
            })
        except ContactMessage.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Сообщение не найдено'}, status=404)
    
    return JsonResponse({'success': False, 'error': 'Неверный метод запроса'}, status=400)
